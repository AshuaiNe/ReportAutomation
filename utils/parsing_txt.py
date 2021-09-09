import re
import openpyxl
import pandas
import xlwt
import string
from pathlib import Path


class ParsingTxt:
    def __init__(self):
      pass

    def get_txt(self, filename):
        txt_json = {}
        line_list = []
        file = open(filename, 'r')
        alphabet = list(string.ascii_uppercase) + [letter1+letter2 for letter1 in string.ascii_uppercase for letter2 in string.ascii_uppercase]
        for x, line in enumerate(file):
            x += 1
            tmp_list = []
            line = line.split("|")
            key = line[0] + '|' + line[1] + '|' + line[2] + '|' + line[3] + '|'
            for cell, value in enumerate(line):
                cell = str(alphabet[cell]) + str(x)
                tmp_list.append({cell: value})
            txt_json[key] = tmp_list
            line_list.append(line)
        file.close()
        df = pandas.DataFrame(line_list)
        filename = str(filename).split('.')[0] + '.xlsx'
        if 'result' in filename:
            challenger = filename.replace('result', 'challenger').replace('copyfile', '')
            df.to_excel(challenger, index=None, sheet_name=Path(str(filename).replace('copyfile', '')).stem)
            df.to_excel(filename, index=None, sheet_name=Path(filename).stem)
        else:
            df.to_excel(filename, index=None, sheet_name=Path(filename).stem)
        return txt_json

    def set_json_to_excel(self, filename, compare_result):
        added_line = []
        removed_line = []
        changed = []
        values_changed = []
        sheet = pandas.ExcelFile(filename, engine='openpyxl').sheet_names[0]
        for key in compare_result.keys():
            if key == 'dictionary_item_added':
                for x in range(len(compare_result[key])):
                    find_str = re.findall(r"\'.*?\'", compare_result[key][x])
                    y = find_str[0].replace("'", "")
                    added_line.append(f'{y}')
            elif key == 'dictionary_item_removed':
                for x in range(len(compare_result[key])):
                    find_str = re.findall(r"\'.*?\'", compare_result[key][x])
                    y = find_str[0].replace("'", "")
                    removed_line.append(f'{y}')
            elif key == 'values_changed':
                for key_1 in compare_result[key]:
                    find_str = re.findall(r"\[.*?\]", key_1)
                    find_str_0 = find_str[2].replace("[", "").replace("]", "").replace("'", "")
                    changed.append(f'{sheet}：{find_str_0}')
                    values_changed.append(compare_result[key][key_1])
            else:
                continue
        dict_frame = {"新增行": added_line, "删除行": removed_line, "差异行": changed, "差异结果": values_changed}
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '比对结果'
        for x, key in enumerate(dict_frame):
            ws.cell(row=1, column=x+1, value=str(key)).value
            for y in range(len(dict_frame[key])):
                ws.cell(row=y+2, column=x+1, value=str(dict_frame[key][y])).value
        wb.save(filename)
